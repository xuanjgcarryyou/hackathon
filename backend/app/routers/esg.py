import hashlib
import json
import uuid
from fastapi import APIRouter, Depends, HTTPException
from pydantic import BaseModel
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select

from app.database import get_db
from app.deps import get_current_user
from app.models.container_batch import ContainerBatch
from app.models.esg_summary import ESGSummary
from app.services.esg_agent import generate_esg_report
from app.services.container_calc import calc_co2e_saved, DISPOSABLE_CO2E_SOURCE, CIRCULAR_CO2E_SOURCE

router = APIRouter()

PACKAGING_KG_PER_UNIT = 0.025


class ESGRequest(BaseModel):
    periodStart: str
    periodEnd: str


@router.post("/esg/generate")
async def create_esg_report(body: ESGRequest, db: AsyncSession = Depends(get_db), user: dict = Depends(get_current_user)):
    company_id = user.get("company_id")
    if not company_id:
        raise HTTPException(status_code=401, detail="company_id missing from token")

    result = await db.execute(
        select(ContainerBatch).where(
            ContainerBatch.company_id == company_id,
            ContainerBatch.status == "collected",
        )
    )
    batches = result.scalars().all()

    total_meals = sum(b.quantity for b in batches)
    circular_meals = sum(b.collected_count or 0 for b in batches)
    co2e_saved = sum(calc_co2e_saved(b.collected_count or 0, 0.15) for b in batches)
    reduced_packaging_kg = circular_meals * PACKAGING_KG_PER_UNIT

    batch_ids = [b.id for b in batches]
    raw_for_hash = [{"id": b.id, "qty": b.quantity, "collected": b.collected_count or 0} for b in batches]
    data_hash = hashlib.sha256(json.dumps(raw_for_hash, sort_keys=True).encode()).hexdigest()
    carbon_factor_source = f"一次性包材：{DISPOSABLE_CO2E_SOURCE}；循環容器：{CIRCULAR_CO2E_SOURCE}"

    report_data = {
        "period_start": body.periodStart,
        "period_end": body.periodEnd,
        "total_meals": total_meals,
        "circular_meals": circular_meals,
        "return_rate": circular_meals / total_meals if total_meals > 0 else 0,
        "reduced_packaging_kg": round(reduced_packaging_kg, 2),
        "co2e_saved": round(co2e_saved, 2),
        "vendor_name": "Loopick",
        "carbon_factor": 0.15,
        "carbon_factor_source": carbon_factor_source,
    }

    report_text_zh, report_text_en, tables, is_fallback = await generate_esg_report(report_data)

    summary = ESGSummary(
        id=str(uuid.uuid4()),
        company_id=company_id,
        period_start=body.periodStart,
        period_end=body.periodEnd,
        total_meals=total_meals,
        circular_meals=circular_meals,
        reduced_packaging_kg=reduced_packaging_kg,
        co2e_saved=co2e_saved,
        report_text_zh=report_text_zh,
        report_text_en=report_text_en,
        tables=tables,
        carbon_factor_source=carbon_factor_source,
        data_hash=data_hash,
        batch_ids=batch_ids,
    )
    db.add(summary)
    await db.commit()

    return {
        "reportId": summary.id,
        "totalMeals": total_meals,
        "circularMeals": circular_meals,
        "reducedPackagingKg": round(reduced_packaging_kg, 2),
        "co2eSaved": round(co2e_saved, 2),
        "reportTextZh": report_text_zh,
        "reportTextEn": report_text_en,
        "tables": tables,
        "carbonFactorSource": carbon_factor_source,
        "dataHash": data_hash,
        "isFallback": is_fallback,
    }


@router.get("/esg/{report_id}/export")
async def export_esg_report(report_id: str, db: AsyncSession = Depends(get_db), user: dict = Depends(get_current_user)):
    company_id = user.get("company_id")
    if not company_id:
        raise HTTPException(status_code=401, detail="company_id missing from token")

    result = await db.execute(
        select(ESGSummary).where(
            ESGSummary.id == report_id,
            ESGSummary.company_id == company_id,
        )
    )
    summary = result.scalar_one_or_none()
    if not summary:
        raise HTTPException(status_code=404, detail="Report not found")

    return_rate = summary.circular_meals / summary.total_meals if summary.total_meals > 0 else 0

    payload = {
        "schemaVersion": "1.0",
        "framework": "GHG Protocol Scope 3 Category 1 — Avoided Emissions",
        "reportId": summary.id,
        "companyId": summary.company_id,
        "period": {"start": summary.period_start, "end": summary.period_end},
        "generatedAt": summary.generated_at.isoformat() if summary.generated_at else None,
        "metrics": {
            "totalMeals": summary.total_meals,
            "circularMeals": summary.circular_meals,
            "returnRate": round(return_rate, 4),
            "reducedPackagingKg": round(summary.reduced_packaging_kg, 3),
            "co2eSavedKg": round(summary.co2e_saved, 3),
        },
        "methodology": {
            "carbonFactorSource": summary.carbon_factor_source,
            "dataHash": summary.data_hash,
            "batchIds": summary.batch_ids,
        },
        "reportText": {
            "zh": summary.report_text_zh,
            "en": summary.report_text_en,
        },
        "tables": summary.tables,
    }

    from fastapi.responses import JSONResponse
    return JSONResponse(
        content=payload,
        headers={"Content-Disposition": f'attachment; filename="esg-report-{report_id[:8]}.json"'},
    )


@router.get("/esg/{report_id}/export/xlsx")
async def export_esg_report_xlsx(report_id: str, db: AsyncSession = Depends(get_db), user: dict = Depends(get_current_user)):
    company_id = user.get("company_id")
    if not company_id:
        raise HTTPException(status_code=401, detail="company_id missing from token")

    result = await db.execute(
        select(ESGSummary).where(
            ESGSummary.id == report_id,
            ESGSummary.company_id == company_id,
        )
    )
    summary = result.scalar_one_or_none()
    if not summary:
        raise HTTPException(status_code=404, detail="Report not found")

    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from fastapi.responses import StreamingResponse

    wb = Workbook()

    # ── 色票 ──
    GREEN = "1A6B3C"
    LIGHT_GREEN = "E8F5E9"
    HEADER_FILL = PatternFill("solid", fgColor=GREEN)
    ROW_FILL = PatternFill("solid", fgColor=LIGHT_GREEN)
    HEADER_FONT = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    BOLD = Font(bold=True, name="Calibri", size=11)
    NORMAL = Font(name="Calibri", size=11)
    CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="CCCCCC")
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hdr(ws, row, col, val):
        c = ws.cell(row=row, column=col, value=val)
        c.font = HEADER_FONT; c.fill = HEADER_FILL
        c.alignment = CENTER; c.border = BORDER

    def cell(ws, row, col, val, bold=False, align=LEFT):
        c = ws.cell(row=row, column=col, value=val)
        c.font = BOLD if bold else NORMAL
        c.alignment = align; c.border = BORDER
        return c

    return_rate = summary.circular_meals / summary.total_meals if summary.total_meals > 0 else 0

    # ── Sheet 1：報告摘要 ──
    ws1 = wb.active
    ws1.title = "報告摘要"
    ws1.column_dimensions["A"].width = 32
    ws1.column_dimensions["B"].width = 28

    hdr(ws1, 1, 1, "項目"); hdr(ws1, 1, 2, "數值")
    rows_s1 = [
        ("報告 ID", summary.id),
        ("企業 ID", summary.company_id),
        ("報告期間", f"{summary.period_start} ～ {summary.period_end}"),
        ("總訂餐份數", summary.total_meals),
        ("循環容器份數", summary.circular_meals),
        ("容器回收率", f"{return_rate * 100:.1f}%"),
        ("減少包材重量 (kg)", round(summary.reduced_packaging_kg, 3)),
        ("CO₂e 避免排放量 (kg)", round(summary.co2e_saved, 3)),
        ("等效植樹數（棵）", round(summary.co2e_saved / 21.8)),
        ("生成時間", str(summary.generated_at)[:19] if summary.generated_at else ""),
    ]
    for i, (k, v) in enumerate(rows_s1, start=2):
        cell(ws1, i, 1, k, bold=True)
        c = cell(ws1, i, 2, v, align=CENTER)
        if i % 2 == 0:
            c.fill = ROW_FILL
            ws1.cell(row=i, column=1).fill = ROW_FILL

    # ── Sheet 2：GHG 清冊 ──
    ws2 = wb.create_sheet("GHG 清冊")
    for col, w in zip("ABCDE", [12, 14, 50, 22, 16]):
        ws2.column_dimensions[chr(ord("A") + "ABCDE".index(col))].width = w
    for col, label in enumerate(["範疇", "類別", "說明", "排放量 (kg CO₂e)", "狀態"], start=1):
        hdr(ws2, 1, col, label)
    ghg_rows = [
        ("Scope 1", "—", "直接排放（燃料燃燒、逸散）", "0", "不適用"),
        ("Scope 2", "—", "電力間接排放", "—", "未追蹤"),
        ("Scope 3", "Cat.1", "包材碳排避免量（Avoided Emissions，循環容器替代一次性包材）", f"{summary.co2e_saved:.2f}", "✓ 已量測"),
        ("Scope 3", "Cat.1", "食材上游完整碳排", "—", "未追蹤"),
        ("Scope 3", "Cat.4", "上游運輸與配送", "—", "未追蹤"),
        ("Scope 3", "Cat.5", "廢棄物處理", "—", "未追蹤"),
    ]
    for i, row in enumerate(ghg_rows, start=2):
        for j, v in enumerate(row, start=1):
            c = cell(ws2, i, j, v, align=CENTER if j in (1, 2, 4, 5) else LEFT)
            if i % 2 == 0:
                c.fill = ROW_FILL

    # ── Sheet 3：中文報告 ──
    ws3 = wb.create_sheet("中文報告")
    ws3.column_dimensions["A"].width = 100
    ws3.row_dimensions[1].height = 22
    hdr(ws3, 1, 1, "ESG 報告（繁體中文）")
    c = ws3.cell(row=2, column=1, value=summary.report_text_zh)
    c.font = NORMAL; c.alignment = Alignment(wrap_text=True, vertical="top")
    ws3.row_dimensions[2].height = 300

    # ── Sheet 4：英文報告 ──
    ws4 = wb.create_sheet("English Report")
    ws4.column_dimensions["A"].width = 100
    ws4.row_dimensions[1].height = 22
    hdr(ws4, 1, 1, "ESG Report (English)")
    c = ws4.cell(row=2, column=1, value=summary.report_text_en)
    c.font = NORMAL; c.alignment = Alignment(wrap_text=True, vertical="top")
    ws4.row_dimensions[2].height = 300

    # ── Sheet 5：方法說明 ──
    ws5 = wb.create_sheet("方法說明")
    ws5.column_dimensions["A"].width = 28
    ws5.column_dimensions["B"].width = 80
    hdr(ws5, 1, 1, "欄位"); hdr(ws5, 1, 2, "說明")
    meta = [
        ("框架", "GHG Protocol Scope 3 Category 1 — Avoided Emissions"),
        ("碳因子來源", summary.carbon_factor_source),
        ("資料稽核碼 (SHA-256)", summary.data_hash),
        ("涵蓋批次數", str(len(summary.batch_ids or []))),
        ("批次 ID 清單", ", ".join(summary.batch_ids or [])),
    ]
    for i, (k, v) in enumerate(meta, start=2):
        cell(ws5, i, 1, k, bold=True)
        cell(ws5, i, 2, v)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"esg-report-{report_id[:8]}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
